#!/usr/bin/env python3
"""
Recalculate all formulas in an Excel file using LibreOffice and report any errors.

Usage:
    python recalc.py <excel_file> [timeout_seconds]

Returns JSON:
    {
      "status": "success" | "errors_found",
      "total_formulas": 42,
      "total_errors": 0,
      "error_summary": {
        "#REF!": {"count": 2, "locations": ["Sheet1!B5", "Sheet1!C10"]}
      }
    }
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

MACRO_CONTENT = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''


def _macro_dir():
    if platform.system() == "Darwin":
        return Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    return Path.home() / ".config/libreoffice/4/user/basic/Standard"


def _ensure_macro():
    macro_file = _macro_dir() / "Module1.xba"
    if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text():
        return True
    if not macro_file.parent.exists():
        subprocess.run(["soffice", "--headless", "--terminate_after_init"],
                       capture_output=True, timeout=15)
        macro_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        macro_file.write_text(MACRO_CONTENT)
        return True
    except Exception:
        return False


def _timeout_cmd(timeout):
    """Return a prefix command for timeout, platform-aware."""
    if platform.system() == "Windows":
        return []
    if platform.system() == "Linux":
        return ["timeout", str(timeout)]
    # macOS: use gtimeout if available, else no timeout wrapper
    try:
        subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1)
        return ["gtimeout", str(timeout)]
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return []


def recalc(filename, timeout=30):
    path = Path(filename)
    if not path.exists():
        return {"error": f"File not found: {filename}"}

    if not _ensure_macro():
        return {"error": "Failed to set up LibreOffice macro"}

    cmd = _timeout_cmd(timeout) + [
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(path.absolute()),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode not in (0, 124):
        return {"error": result.stderr or "LibreOffice recalculation failed"}

    # Scan for errors in recalculated file
    try:
        wb = load_workbook(filename, data_only=True)
        error_map = {e: [] for e in EXCEL_ERRORS}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for err in EXCEL_ERRORS:
                            if err in cell.value:
                                error_map[err].append(f"{sheet}!{cell.coordinate}")
                                break
        wb.close()

        total_errors = sum(len(v) for v in error_map.values())
        summary = {
            err: {"count": len(locs), "locations": locs[:20]}
            for err, locs in error_map.items() if locs
        }

        # Count formulas
        wb2 = load_workbook(filename, data_only=False)
        formula_count = sum(
            1
            for sheet in wb2.sheetnames
            for row in wb2[sheet].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        wb2.close()

        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_formulas": formula_count,
            "total_errors": total_errors,
            **({"error_summary": summary} if summary else {}),
        }
    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc(sys.argv[1], timeout), indent=2))


if __name__ == "__main__":
    main()
